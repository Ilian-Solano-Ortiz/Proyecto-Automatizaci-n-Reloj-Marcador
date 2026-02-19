import pandas as pd
import os, tempfile, re
import win32com.client as win32
from openpyxl import load_workbook
from tkinter import messagebox



#Convierte valores numéricos o texto a formato HH:MM
def convertir_hora_excel(valor):
    try:
        if pd.isna(valor) or valor == "":
            return None
        if isinstance(valor, str):
            s = valor.strip()
            if re.match(r"^\d{1,2}:\d{2}$", s):
                return s
            return None
        if isinstance(valor, (int, float)):
            horas = int(valor * 24)
            minutos = int(round((valor * 24 - horas) * 60))
            return f"{horas:02d}:{minutos:02d}"
    except:
        return None
    return None

#Lee el formato del 'Reporte de Eventos de Asistencia' del MEP.
# - Detecta múltiples horas en una celda y usa la primera como entrada y la última como salida.
#- Incluye también a los profesores que no registraron asistencia en la semana,
#  mostrando las fechas del periodo con horarios vacíos.
def leer_excel_global(ruta_excel):
    wb = load_workbook(filename=ruta_excel, data_only=True)
    ws = wb.active
    data = [[str(c).strip() if c else "" for c in row] for row in ws.iter_rows(values_only=True)]
    registros = []
    fechas = []
    profesores = []

    
    # Buscar rango de periodo (ej: 2025-03-03 ~ 2025-03-10)
    fechas=buscar_rango_periodo(data)
    if not fechas:
        fechas = [f"Día {i+1}" for i in range(10)]

    # Recorre archivo buscando bloques "ID:" con sus filas de horas
    recorrer_buscando_id(data, profesores,fechas, registros)
    df = pd.DataFrame(registros)
    
    if df.empty:
        from tkinter import messagebox
        messagebox.showerror("Error", "No se pudieron extraer registros válidos del archivo.")
        return pd.DataFrame()

    df = df.fillna("")
    print(f"Se cargaron {len(df)} registros (incluyendo profesores sin marcaciones).")
    return df

def buscar_rango_periodo(data):
    for fila in data:
        texto = " ".join(fila)
        if "periodo" in texto.lower() and "~" in texto:
            match = re.search(r"(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})", texto)
            if match:
                inicio, fin = match.groups()
                fechas = pd.date_range(inicio, fin).strftime("%Y-%m-%d").tolist()
                break
    return fechas;

def recorrer_buscando_id(data, profesores,fechas, registros):
     for i, fila in enumerate(data):
        fila_texto = " ".join(fila).lower()
        if "id:" in fila_texto and "nombre:" in fila_texto:
            # Extraer datos del profesor
            id_match = re.search(r"id[:\s]*([0-9]+)", fila_texto)
            nombre_match = re.search(r"nombre[:\s]*(.*?)\s*departamento[:\s]*", fila_texto, re.IGNORECASE)
            depto_match = re.search(r"departamento[:\s]*([A-Za-z\s]+)", fila_texto, re.IGNORECASE)

            if not id_match or not nombre_match:
                continue

            id_actual = id_match.group(1).strip()
            nombre_actual = nombre_match.group(1).strip().title().replace("  ", " ")
            depto_actual = depto_match.group(1).strip().title() if depto_match else ""

            profesores.append((id_actual, nombre_actual, depto_actual))

            # Lee filas siguientes (entradas/salidas)
            fila_entradas = data[i + 1] if i + 1 < len(data) else []
            fila_salidas = data[i + 2] if i + 2 < len(data) else []

            hubo_registro = False

            for j, fecha in enumerate(fechas):
                celda = ""
                if j < len(fila_entradas):
                    celda = fila_entradas[j]
                elif j < len(fila_salidas):
                    celda = fila_salidas[j]

                horas = extraer_horas(celda)
                if not horas:
                    continue

                hubo_registro = True
                entrada = horas[0]
                salida = horas[-1] if len(horas) > 1 else ""

                registros.append({
                    "ID": id_actual,
                    "Nombre": nombre_actual,
                    "Departamento": depto_actual,
                    "Fecha": fecha,
                    "Entrada": entrada,
                    "Salida": salida
                })

            # Si no hubo registros, lo agregamos igual con todas las fechas vacías
            if not hubo_registro:
                for fecha in fechas:
                    registros.append({
                        "ID": id_actual,
                        "Nombre": nombre_actual,
                        "Departamento": depto_actual,
                        "Fecha": fecha,
                        "Entrada": "",
                        "Salida": ""
                    })

def extraer_horas(celda):
    if not celda:
          return []
    return re.findall(r"(\d{1,2}:\d{2})", str(celda))
